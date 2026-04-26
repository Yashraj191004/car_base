import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill

EXCEL_FILE = "vehicle_oils.xlsx"

# -----------------------------
# SETUP EXCEL WORKBOOK
# -----------------------------
def setup_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehicle Oils"

    # Define headers (matching the DB columns)
    headers = [
        "Source File", "Year", "Make", "Model", "Engine", 
        "Oil Type", "Recommendation Level", "Temperature",
        "Capacity With Filter (Quarts)", "Capacity With Filter (Liters)",
        "Capacity Without Filter (Quarts)", "Capacity Without Filter (Liters)"
    ]

    # Write headers
    ws.append(headers)

    # Style headers: Bold, light gray background, frozen top row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    ws.freeze_panes = "A2"
    
    return wb, ws


# -----------------------------
# INSERT FLATTENED DATA
# -----------------------------
def insert_flat_data(ws, filename, vehicle_info, engines):
    year = vehicle_info.get("year")
    make = vehicle_info.get("make")
    model = vehicle_info.get("model")

    # Skip files with no engine data.
    if not engines:
        return

    for engine_name, engine_data in engines.items():
        capacity = engine_data.get("oil_capacity", {})
        with_filter = capacity.get("with_filter", {}) or {}
        without_filter = capacity.get("without_filter", {}) or {}

        oil_recs = engine_data.get("oil_recommendations", [])

        # If an engine exists but has no recommendations, still write one row.
        if not oil_recs:
            ws.append([
                filename,
                year,
                make,
                model,
                engine_name,
                "N/A",
                "N/A",
                "N/A",
                with_filter.get("quarts"),
                with_filter.get("liters"),
                without_filter.get("quarts"),
                without_filter.get("liters")
            ])
            continue

        for rec in oil_recs:
            oil_type = rec.get("oil_type")
            level = rec.get("recommendation_level")
            temps = rec.get("temperature_condition", [])

            # If no temperature, still insert one row
            if not temps:
                temps = ["N/A"]

            for temp in temps:
                row_data = [
                    filename,
                    year,
                    make,
                    model,
                    engine_name,
                    oil_type,
                    level,
                    temp,
                    with_filter.get("quarts"),
                    with_filter.get("liters"),
                    without_filter.get("quarts"),
                    without_filter.get("liters")
                ]
                ws.append(row_data)


# -----------------------------
# MAIN MIGRATION FUNCTION
# -----------------------------
def migrate_json_to_excel(json_file):
    # Load JSON
    try:
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        print("✗ JSON file not found")
        return
    except json.JSONDecodeError:
        print("✗ Invalid JSON format")
        return

    # Initialize Excel Workbook
    wb, ws = setup_workbook()

    try:
        for filename, vehicle_data in data.items():
            print(f"Processing: {filename}")

            vehicle_info = vehicle_data.get("Vehicle", {})
            engines = vehicle_data.get("engines", {})

            insert_flat_data(ws, filename, vehicle_info, engines)

        # Optional: Auto-adjust column widths for better readability
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = min(adjusted_width, 40) # Cap width at 40

        # Save Workbook
        wb.save(EXCEL_FILE)
        print("\n✓ Data inserted successfully!")
        print(f"✓ Excel file saved as: {EXCEL_FILE}")

    except Exception as e:
        print("✗ Error:", e)


# -----------------------------
# RUN SCRIPT
# -----------------------------
if __name__ == "__main__":
    print("=" * 50)
    print("Vehicle Oil JSON to Excel Converter")
    print("=" * 50)

    # Auto-detect JSON file
    possible_paths = [
        "structured_results.json",
        Path(__file__).parent / "structured_results.json"
    ]

    json_file = None
    for path in possible_paths:
        if Path(path).exists():
            json_file = str(path)
            break

    if not json_file:
        print("✗ structured_results.json not found")
    else:
        print(f"✓ Using: {json_file}")
        migrate_json_to_excel(json_file)